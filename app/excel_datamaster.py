from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Dict

import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


LISTED_FACILITIES = {
    "BANK GARANSI",
    "KREDIT LOKAL",
    "TIME LOAN",
    "KREDIT MULTI FASILITAS",
    "L/C",
    "TRUST RECEIPT",
}

COLS = ["Nama", "ACC No", "Cabang", "Fasilitas", "Plafond", "Tgl PMK/PK Akhir", "IsPriority", "SumberFile"]


def _ensure_cols(df: pd.DataFrame) -> pd.DataFrame:
    for c in COLS:
        if c not in df.columns:
            df[c] = None
    df = df[COLS].copy()
    df["ACC No"] = df["ACC No"].astype(str)  # keep leading zero
    return df


def _normalize_date(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Tgl PMK/PK Akhir"] = pd.to_datetime(df["Tgl PMK/PK Akhir"], errors="coerce").dt.date
    return df


def _split_priority(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Split DEBITUR-level sesuai rule:
    - Debitur yang punya >=1 fasilitas prioritas -> masuk PRIORITAS, TAPI hanya baris prioritasnya yang ditampilkan.
    - Debitur tanpa fasilitas prioritas sama sekali -> masuk NON, tampilkan semua barisnya.
    """
    d = df.copy()

    # Pakai flag dari engine (future-proof)
    d["IsPriority"] = d["IsPriority"].astype(bool)

    has_pri = d.groupby("Nama")["IsPriority"].any().rename("HasPriority")
    d = d.merge(has_pri, on="Nama", how="left")

    pri = d[(d["HasPriority"] == True) & (d["IsPriority"] == True)].copy()
    non = d[d["HasPriority"] == False].copy()

    pri = pri.drop(columns=["HasPriority"])
    non = non.drop(columns=["HasPriority"])
    return pri, non


def _sort_block_style(df: pd.DataFrame) -> pd.DataFrame:
    """
    ATURAN FINAL (BLOCK STABLE):
    - Urutan utama: debitur diurut berdasarkan tanggal paling awal per debitur.
    - Debitur harus 1 blok utuh (tidak boleh tercampur) -> pakai Nama sebagai tie-breaker.
    - Dalam debitur: urut berdasarkan tanggal fasilitas.
    """
    df = df.copy()
    df["Tgl PMK/PK Akhir"] = pd.to_datetime(df["Tgl PMK/PK Akhir"], errors="coerce").dt.date

    key = df.groupby("Nama")["Tgl PMK/PK Akhir"].min().rename("Debitur_Earliest")
    df = df.merge(key, on="Nama", how="left")

    df = df.sort_values(
        ["Debitur_Earliest", "Nama", "Tgl PMK/PK Akhir", "Fasilitas", "ACC No"],
        na_position="last",
        kind="mergesort",
    )
    return df.drop(columns=["Debitur_Earliest"])

def _write_df_sheet(ws, df: pd.DataFrame, header_fill, header_font, center, widths: dict):
    # header
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 18)

    # rows
    for r_i, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(df.columns, start=1):
            ws.cell(row=r_i, column=c, value=row[col])

    ws.freeze_panes = "A2"


def build_outputs(raw: pd.DataFrame, out_dir: Path, job_id: str) -> Dict[str, Path]:
    """
    Outputs:
      - output/FULL/{job_id}_gabungan.xlsx
      - output/FULL/{job_id}_per-bulan.zip  (tetap jika sudah ada logic zip Anda sebelumnya)
    """
    out_dir = Path(out_dir)
    full_dir = out_dir / "FULL"
    full_dir.mkdir(parents=True, exist_ok=True)

    # --- prepare base df ---
    df = _ensure_cols(raw)
    df = _normalize_date(df)

    # fasilitas uppercase
    df["Fasilitas"] = df["Fasilitas"].astype(str).str.upper().str.replace(r"\s+", " ", regex=True)

    pri, non = _split_priority(df)
    pri = _sort_block_style(pri)
    non = _sort_block_style(non)

    # --- workbook styles ---
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    title_fill = PatternFill("solid", fgColor="E6F0FF")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    widths = {
        "Nama": 42,
        "ACC No": 16,
        "Cabang": 8,
        "Fasilitas": 26,
        "Plafond": 16,
        "Tgl PMK/PK Akhir": 16,
        "SumberFile": 18,
    }

    # --- build Gabungan sheet in "v2 style" ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Gabungan"

    title_pri = "PRIORITAS (BG/KREDIT LOKAL/TIME LOAN/KREDIT MULTI FASILITAS/L/C/TRUST RECEIPT)"
    title_non = "NON-PRIORITAS (Nama tanpa fasilitas prioritas)"

    def write_title(row: int, text: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        cell.alignment = left
        cell.fill = title_fill

    def write_header(row: int):
        hdr = ["Nama", "ACC No", "Cabang", "Fasilitas", "Plafond", "Tgl PMK/PK Akhir"]
        for c, h in enumerate(hdr, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 18)

    r = 1
    write_title(r, title_pri); r += 1
    write_header(r); r += 1

    pri_view = pri[["Nama", "ACC No", "Cabang", "Fasilitas", "Plafond", "Tgl PMK/PK Akhir"]].copy()
    for _, row in pri_view.iterrows():
        ws.cell(row=r, column=1, value=row["Nama"])
        ws.cell(row=r, column=2, value=row["ACC No"])
        ws.cell(row=r, column=3, value=row["Cabang"])
        ws.cell(row=r, column=4, value=row["Fasilitas"])
        ws.cell(row=r, column=5, value=row["Plafond"])
        ws.cell(row=r, column=6, value=row["Tgl PMK/PK Akhir"])
        r += 1

    r += 1  # blank row
    write_title(r, title_non); r += 1
    write_header(r); r += 1

    non_view = non[["Nama", "ACC No", "Cabang", "Fasilitas", "Plafond", "Tgl PMK/PK Akhir"]].copy()
    for _, row in non_view.iterrows():
        ws.cell(row=r, column=1, value=row["Nama"])
        ws.cell(row=r, column=2, value=row["ACC No"])
        ws.cell(row=r, column=3, value=row["Cabang"])
        ws.cell(row=r, column=4, value=row["Fasilitas"])
        ws.cell(row=r, column=5, value=row["Plafond"])
        ws.cell(row=r, column=6, value=row["Tgl PMK/PK Akhir"])
        r += 1

    ws.freeze_panes = "A3"

    # --- other sheets (audit) ---
    ws_pri = wb.create_sheet("Prioritas_Detail")
    _write_df_sheet(ws_pri, pri[COLS], header_fill, header_font, center, widths)

    ws_non = wb.create_sheet("NonPrioritas_Detail")
    _write_df_sheet(ws_non, non[COLS], header_fill, header_font, center, widths)

    ws_raw = wb.create_sheet("Raw_All")
    _write_df_sheet(ws_raw, df[COLS], header_fill, header_font, center, widths)

    # --- save full ---
    full_path = full_dir / f"{job_id}_gabungan.xlsx"
    wb.save(full_path)

    # ZIP per bulan:
    # Jika Anda sudah punya logic zip sebelumnya, pertahankan. Untuk sekarang, buat placeholder file zip agar flow tidak rusak.
    zip_path = full_dir / f"{job_id}_per-bulan.zip"
    if not zip_path.exists():
        zip_path.write_bytes(b"")  # placeholder

    return {"full": full_path, "zip": zip_path}
